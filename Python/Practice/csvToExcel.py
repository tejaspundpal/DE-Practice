import pandas as pd
from openpyxl import load_workbook
from pandas.api.types import is_string_dtype
from openpyxl.styles import Font
 
def csv_to_excel_with_format(input_csv, output_excel,delimiter):
    with open(input_csv, 'r', encoding='utf8') as file:
            df = pd.read_csv(file,delimiter=delimiter) 

    df.columns = [col.replace('_', ' ').strip() for col in df.columns]
 
    # Function to check if a column contains date-like values
    def is_date_column(series):
        if is_string_dtype(series) or pd.api.types.is_object_dtype(series):
            try:
                pd.to_datetime(series.dropna().iloc[:5])  # Check first 5 non-NA values
                return True
            except (ValueError, TypeError):
                return False
        return False
 
    # Convert only date-like columns to datetime
    for col in df.columns:
        if is_date_column(df[col]):
            df[col] = pd.to_datetime(df[col], errors='coerce')  # Safely coerce invalid dates to NaT
 
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Goal Details')
 
        # Format date columns in the Excel file
        workbook = writer.book
        worksheet = writer.sheets['Goal Details']
 
        for col_idx, col_name in enumerate(df.columns, 1):  # OpenPyXL columns are 1-indexed
            header_cell = worksheet.cell(row=1, column=col_idx)
            header_cell.font = Font(bold=False)
            if pd.api.types.is_datetime64_any_dtype(df[col_name]):
                for row_idx in range(2, len(df) + 2):  # Skip the header row
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell.number_format = 'MMM D, YY' 
 
if __name__ == "__main__":
    input_csv = 'promo_varied_11.csv'  # Replace with your actual CSV file path
    output_excel = 'output.xlsx'  # Save the output file in the current directory
    delimiter = "~"
    try:
        csv_to_excel_with_format(input_csv, output_excel,delimiter)
        print("CSV successfully converted to Excel with formatted dates. File saved as 'output.xlsx'.")
    except PermissionError:
        print("PermissionError: Please close the output file if it is open and try again.")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")